# -*- coding: utf-8 -*-
"""
patch_all.py — 슈퍼로봇대전 J  전체 챕터 한글 대사 삽입 패치

특징
----
* 번역 매칭 엑셀이 있는 챕터를 한 번에 모두 패치한다.
  (엑셀이 1개뿐이면 1화만 패치 — patch_ep1.py 와 동일 동작)
* 새 한글 블록은 ROM 빈 공간부터 차례로 기록하고,
  공간이 모자라면 ROM 파일을 자동으로 확장한다(최대 32MB).
* 희생슬롯 우선순위는 '이번에 번역 안 하는 챕터'의 사용량만 보고
  동적으로 계산 → 번역 챕터가 많아질수록 더 안전해진다.

엑셀 파일 이름
--------------
    srwj_matched_block_NNN.xlsx
        NNN = 대사 블록 순번(1-based). 001 = 1화.
    번역한 챕터의 엑셀만 폴더에 두면 된다(나머지는 일본어 유지).

사용법
------
    python patch_all.py
    python patch_all.py --rom kr.gba --xlsx-dir ./translations \\
                        --out srwj_korean.gba --reserve 7
"""

import argparse
import glob
import os
import re
import struct
import sys

import srwj_decode as D
import srwj_parser as P
from srwj_codec import HangulCodec
from srwj_wrap import fit_turn_lines
from srwj_inject_lib import (compute_victim_rank, build_korean_block,
                             ensure_rom_capacity)

FREE_SPACE_ADDR = 0xFB2400      # 새 블록을 쌓기 시작할 빈 공간


# ──────────────────────────────────────────────────────────
def find_one(patterns, label):
    for pat in patterns:
        hits = sorted(glob.glob(pat))
        if hits:
            return hits[0]
    sys.exit(f'[오류] {label} 을(를) 찾을 수 없습니다: {patterns}')


def load_excel_kr(path):
    """매칭 엑셀에서 턴순번 → (한국어, 화자) 정보를 읽는다.

    Returns:
        (kr_map, speaker_map)
          kr_map     : {턴번호: 한국어 텍스트}
          speaker_map: {턴번호: 화자 이름(번역)}  (첫 줄 폭 계산용)
    """
    from openpyxl import load_workbook
    ws = load_workbook(path, data_only=True)['매칭 결과']
    kr = {}
    spk = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 8).value          # 8번 열 = 한국어(번역TXT)
        s = ws.cell(r, 5).value          # 5번 열 = 화자(번역)
        if v is not None and str(v).strip():
            kr[r - 2] = v
        # 화자는 한국어 유무와 무관하게 기록 (빈 화자=독백도 표시)
        spk[r - 2] = ('' if s is None else str(s).strip())
    return kr, spk


# ──────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description='슈로대J 전체 챕터 한글 삽입')
    ap.add_argument('--rom', help='입력 한글패치 ROM')
    ap.add_argument('--xlsx-dir', default='.', help='매칭 엑셀들이 있는 폴더')
    ap.add_argument('--korea', help='korea2350.txt')
    ap.add_argument('--japan', help='japan2350.txt')
    ap.add_argument('--out', help='출력 ROM')
    ap.add_argument('--reserve', type=int, default=7,
                    help='첫 줄 "화자「" 폭 (기본 7)')
    ap.add_argument('--addr', type=lambda x: int(x, 0), default=FREE_SPACE_ADDR,
                    help='새 블록 기록 시작 주소 (기본 0xFB2400)')
    ap.add_argument('--expand-dict', action='store_true',
                    help='사전 용량(약 1930자) 초과 시 seg1 을 물리적으로 '
                         '확장 (전체/거의 전체 챕터 번역 시 권장)')
    ap.add_argument('--keep-jp', action='store_true',
                    help='번역이 없는 턴을 원본 일본어로 둔다 '
                         '(기본은 "번역 없음" 표시를 넣음)')
    ap.add_argument('--placeholder-text', default='번역 없음',
                    help='미번역 턴에 넣을 표시 문구 (기본 "번역 없음")')
    args = ap.parse_args()

    placeholder = None if args.keep_jp else args.placeholder_text

    here = os.path.dirname(os.path.abspath(__file__))
    rom_path   = args.rom   or find_one([os.path.join(here, '*.gba'), '*.gba'], 'ROM')
    korea_path = args.korea or find_one([os.path.join(here, 'korea2350.txt'), 'korea2350.txt'], 'korea2350.txt')
    japan_path = args.japan or find_one([os.path.join(here, 'japan2350.txt'), 'japan2350.txt'], 'japan2350.txt')
    out_path   = args.out   or os.path.join(here, 'srwj_korean_all.gba')

    print('=' * 64)
    print(' 슈퍼로봇대전 J  전체 챕터 한글 대사 삽입')
    print('=' * 64)

    # ── 1. ROM & 블록 목록 ───────────────────────────────
    rom = bytearray(D.load_rom(rom_path))
    rom_orig_size = len(rom)
    dic = D.Dictionary(bytes(rom))
    idx = list(D.load_archive_index(bytes(rom)))
    blocks = D.find_all_dialogue_blocks(bytes(rom))
    print(f'  ROM        : {rom_path}  ({rom_orig_size:,} 바이트)')
    print(f'  대사 블록  : {len(blocks)}개 '
          f'(archive 인덱스 {blocks[0]["archive_idx"]}~{blocks[-1]["archive_idx"]})')

    # ── 2. 번역 엑셀 수집 ────────────────────────────────
    #  파일명 srwj_matched_block_NNN.xlsx 의 NNN 은 'archive 인덱스'다.
    #  (매칭 도구가 archive 인덱스로 파일명을 붙임)
    by_archive = {meta['archive_idx']: meta for meta in blocks}
    xlsx_map = {}                # archive 인덱스 → 엑셀 경로
    unmatched_files = []
    for path in glob.glob(os.path.join(args.xlsx_dir, '*.xlsx')):
        base = os.path.basename(path)
        if 'summary' in base.lower():
            continue
        m = re.search(r'block[_-]?(\d+)', base, re.I)
        if not m:
            continue
        ai = int(m.group(1))
        if ai in by_archive:
            xlsx_map[ai] = path
        else:
            unmatched_files.append((base, ai))
    if not xlsx_map:
        sys.exit(f'[오류] {args.xlsx_dir} 에서 패치할 수 있는 '
                 f'srwj_matched_block_NNN.xlsx 파일을 찾지 못했습니다.')
    translated_archive = set(xlsx_map)
    print(f'  번역 엑셀  : {len(xlsx_map)}개 '
          f'(archive 인덱스 {sorted(xlsx_map)[:10]}'
          f'{" ..." if len(xlsx_map) > 10 else ""})')
    if unmatched_files:
        print(f'  ⚠ 대사 블록에 없는 archive 인덱스 파일 {len(unmatched_files)}개 '
              f'— 무시됨: {[f"{b}(→{a})" for b, a in unmatched_files[:5]]}')

    # ── 3. 희생슬롯 우선순위 동적 계산 ───────────────────
    print()
    print('  [희생슬롯 우선순위 계산]  (번역 안 하는 챕터 사용량만 집계)')
    victim_rank, code_use = compute_victim_rank(bytes(rom), dic, blocks,
                                                translated_archive)
    free_now = sum(1 for c in victim_rank if code_use.get(c, 0) == 0)
    print(f'   사용량 0 인 seg1 슬롯: {free_now} / 1851')

    # ── 4. 코덱 준비 — 모든 챕터 한국어로 plan() ─────────
    codec = HangulCodec(bytes(rom), korea_path, japan_path, victim_rank,
                        expand_mode=args.expand_dict)
    all_kr = []
    kr_per_block = {}            # archive 인덱스 → {턴번호: 한국어}
    spk_per_block = {}           # archive 인덱스 → {턴번호: 화자}
    for ai, path in sorted(xlsx_map.items()):
        kr, spk = load_excel_kr(path)
        kr_per_block[ai] = kr
        spk_per_block[ai] = spk
        all_kr.extend(kr.values())
    codec.plan(all_kr + ([placeholder] if placeholder else []))
    print()
    print('  [사전 코드 배정]'
          + ('  (확장 모드)' if args.expand_dict else '  (희생슬롯 모드)'))
    print(codec.report())
    if placeholder:
        print(f'   미번역 턴 표시 문구: "{placeholder}"')

    write_addr = (args.addr + 3) & ~3

    if args.expand_dict:
        if codec.n_expansion > 0:
            from srwj_inject_lib import expand_dictionary
            write_addr = expand_dictionary(rom, codec.n_expansion, write_addr)
            print(f'   사전 확장: seg1 +{codec.n_expansion}코드, '
                  f'seg2 테이블 이전 완료')
    else:
        if codec.victim_count > 1851 or codec.unresolved:
            print('   ★ 사전 용량(희생슬롯)이 부족합니다.')
            print('     → --expand-dict 옵션으로 다시 실행하세요.')

    # ── 5. 챕터별 한국어 블록 빌드 + 빈 공간에 배치 ──────
    print()
    print('  [챕터별 빌드 & 배치]')
    tot_kr_turns = tot_warn_major = tot_ph_turns = 0
    patched = 0

    for meta in blocks:
        ai = meta['archive_idx']
        if ai not in translated_archive:
            continue
        new_block, st = build_korean_block(
            bytes(rom), dic, meta, kr_per_block[ai], codec,
            args.reserve, fit_turn_lines, placeholder=placeholder,
            speaker_by_turn=spk_per_block.get(ai))

        end = write_addr + len(new_block)
        ensure_rom_capacity(rom, end)            # 필요 시 ROM 확장
        rom[write_addr:end] = new_block

        # 아카이브 인덱스 repoint
        struct.pack_into('<I', rom, D.IDX_BASE + ai * 4,
                         write_addr - D.IDX_BASE)

        tot_kr_turns += st['kr_turns']
        tot_warn_major += st['warn_major']
        tot_ph_turns += st['ph_turns']
        patched += 1
        notes = []
        if st['ph_turns']:
            notes.append(f"미번역 {st['ph_turns']}")
        if st['warn_major']:
            notes.append(f"매칭점검 {st['warn_major']}")
        flag = ('  ⚠ ' + ', '.join(notes)) if notes else ''
        print(f'   archive {ai:>3}: '
              f'0x{write_addr:08X}  {len(new_block):>6}B  '
              f'한국어 {st["kr_turns"]}턴{flag}')
        write_addr = (end + 3) & ~3

    # ── 6. 사전 패치 적용 ────────────────────────────────
    for off, sjis in codec.dict_patches:
        rom[off:off + 2] = sjis

    # ── 7. 저장 ─────────────────────────────────────────
    with open(out_path, 'wb') as f:
        f.write(rom)

    print()
    print('  [완료]')
    print(f'   패치한 챕터        : {patched}')
    print(f'   한국어 인코딩 턴   : {tot_kr_turns}')
    if placeholder:
        print(f'   미번역 표시 턴     : {tot_ph_turns} ("{placeholder}")')
    if args.expand_dict:
        print(f'   사전 확장 새 코드  : {codec.n_expansion}')
    else:
        print(f'   사전 희생슬롯      : {codec.victim_count}')
    print(f'   매칭 점검 권장 턴  : {tot_warn_major} (줄 수 +2 이상)')
    if len(rom) != rom_orig_size:
        print(f'   ROM 크기 확장      : {rom_orig_size:,} → {len(rom):,} 바이트 '
              f'({len(rom)//1024//1024}MB)')
    else:
        print(f'   ROM 크기           : {len(rom):,} 바이트 (확장 불필요)')
    print(f'   마지막 기록 주소   : 0x{write_addr:08X} '
          f'(빈 공간 끝까지 0x{len(rom):08X})')
    print(f'   ✓ 저장: {out_path}')

    # ── 8. 검증 ─────────────────────────────────────────
    verify(out_path, blocks, translated_archive, codec)


# ──────────────────────────────────────────────────────────
def verify(rom_path, blocks, translated_archive, codec):
    """패치 ROM 을 다시 읽어 번역 챕터들이 정상 파싱되는지 확인."""
    print()
    print('  [검증] 패치 ROM 재파싱')
    rom = D.load_rom(rom_path)
    dic = D.Dictionary(rom)
    idx = D.load_archive_index(rom)
    kanji2ko = {v: k for k, v in codec.ko2kanji.items()}

    ok = fail = 0
    sample_done = False
    for meta in blocks:
        ai = meta['archive_idx']
        if ai not in translated_archive:
            continue
        addr = D.IDX_BASE + idx[ai]
        ptrs = P.read_dialogue_pointers(rom, addr)
        info = P.parse_dialogue_block(rom, addr, ptrs[-1], dic)
        block_ok = all(d['parse_ok'] for d in info['dialogues'])
        if block_ok:
            ok += 1
        else:
            fail += 1
            print(f'   ★ archive {ai} 파싱 실패')
        # 첫 번역 챕터 한 턴을 한글로 복원해 미리보기
        if not sample_done:
            for dlg in info['dialogues']:
                for turn in dlg['turns']:
                    s = ''
                    for line in turn['lines']:
                        for code, w in D.tokenize(line):
                            s += ''.join(kanji2ko.get(c, c)
                                         for c in dic.decode(code))
                    if s.strip():
                        print(f'   미리보기(archive {ai}): {s[:50]}')
                        sample_done = True
                        break
                if sample_done:
                    break
    print(f'   번역 챕터 {ok + fail}개 중 파싱 정상 {ok}, 실패 {fail}')
    if fail == 0:
        print('   ✓ 모든 번역 챕터 구조 정상')


if __name__ == '__main__':
    main()

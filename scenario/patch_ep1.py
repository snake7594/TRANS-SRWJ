# -*- coding: utf-8 -*-
"""
patch_ep1.py — 슈퍼로봇대전 J  1화 한글 대사 삽입 패치 (PC 실행용)

하는 일
--------
1. 한글 패치 ROM 의 1화 대사 블록(블록1)을 파싱한다.
2. 매칭 엑셀(srwj_matched_block_001.xlsx)의 한국어 번역을 읽는다.
3. 각 턴의 한국어를 대사창 폭(25)에 맞춰 줄바꿈하고 게임 코드로 인코딩한다.
4. 새 1화 블록을 만들어 ROM 끝 빈 공간에 기록하고,
   아카이브 인덱스가 그쪽을 가리키도록 포인터를 수정한다.
5. 사전에 없는 한자(한글 글자)는 seg1 의 거의 안 쓰이는 슬롯을 덮어써서
   인코딩 가능하게 만든다.
6. 패치된 ROM 을 저장하고, 다시 읽어 검증한다.

사용법
------
    python patch_ep1.py
        (스크립트와 같은 폴더의 *.gba, *.xlsx, korea2350.txt, japan2350.txt 자동 탐색)

    python patch_ep1.py --rom kr.gba --excel srwj_matched_block_001.xlsx \\
                        --out kr_ep1_patched.gba --reserve 7
"""

import argparse
import glob
import os
import struct
import sys

import srwj_decode as D
import srwj_parser as P
from srwj_codec import HangulCodec, normalize_text
from srwj_wrap import fit_turn_lines

FREE_SPACE_ADDR = 0xFB2400      # ROM 끝 빈 공간 시작 (4바이트 정렬)
BLOCK1_ARCHIVE_IDX = 1          # 1화 = archive1 인덱스 1번


# ──────────────────────────────────────────────────────────
#  유틸
# ──────────────────────────────────────────────────────────
def find_one(patterns, label):
    for pat in patterns:
        hits = sorted(glob.glob(pat))
        if hits:
            return hits[0]
    sys.exit(f'[오류] {label} 파일을 찾을 수 없습니다: {patterns}')


def rebuild_turn_original(turn) -> bytes:
    """파싱된 턴을 원본(일본어) 바이트열로 복원."""
    out = bytearray([turn['cid'], turn['attr']])
    if turn['aux'] is not None:
        out += turn['aux']
    n = len(turn['lines'])
    for i, line in enumerate(turn['lines']):
        lb = (len(line) & 0x7F) | (0x80 if i == n - 1 else 0)
        out.append(lb)
        out += line
    return bytes(out)


def rebuild_turn_korean(turn, kr_lines, codec) -> bytes:
    """파싱된 턴 헤더 + 한국어 줄들로 새 턴 바이트열 생성."""
    out = bytearray([turn['cid'], turn['attr']])
    if turn['aux'] is not None:
        out += turn['aux']
    encoded = [codec.encode_line(s) for s in kr_lines]
    n = len(encoded)
    for i, enc in enumerate(encoded):
        if len(enc) > 0x7F:
            raise ValueError(f'줄 길이 {len(enc)} > 127 바이트 (인코딩 초과)')
        lb = (len(enc) & 0x7F) | (0x80 if i == n - 1 else 0)
        out.append(lb)
        out += enc
    return bytes(out)


# ──────────────────────────────────────────────────────────
#  메인
# ──────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description='슈로대J 1화 한글 대사 삽입')
    ap.add_argument('--rom',   help='입력 한글패치 ROM (.gba)')
    ap.add_argument('--excel', help='매칭 엑셀 (.xlsx)')
    ap.add_argument('--korea', help='korea2350.txt')
    ap.add_argument('--japan', help='japan2350.txt')
    ap.add_argument('--out',   help='출력 ROM 경로')
    ap.add_argument('--reserve', type=int, default=7,
                    help='첫 줄에서 "화자「" 가 차지하는 폭 (기본 7)')
    ap.add_argument('--addr', type=lambda x: int(x, 0), default=FREE_SPACE_ADDR,
                    help='새 블록을 기록할 빈 공간 주소 (기본 0xFB2400)')
    args = ap.parse_args()

    here = os.path.dirname(os.path.abspath(__file__))
    rom_path   = args.rom   or find_one([os.path.join(here, '*.gba'), '*.gba'], 'ROM')
    excel_path = args.excel or find_one([os.path.join(here, '*.xlsx'), '*.xlsx'], '매칭 엑셀')
    korea_path = args.korea or find_one([os.path.join(here, 'korea2350.txt'), 'korea2350.txt'], 'korea2350.txt')
    japan_path = args.japan or find_one([os.path.join(here, 'japan2350.txt'), 'japan2350.txt'], 'japan2350.txt')
    out_path   = args.out   or os.path.join(here, 'srwj_ep1_korean.gba')

    print('=' * 60)
    print(' 슈퍼로봇대전 J  1화 한글 대사 삽입')
    print('=' * 60)
    print(f'  ROM   : {rom_path}')
    print(f'  엑셀  : {excel_path}')
    print(f'  출력  : {out_path}')
    print()

    # ── 1. ROM 로드 & 블록1 파싱 ──────────────────────────
    rom = bytearray(D.load_rom(rom_path))
    dic = D.Dictionary(bytes(rom))
    idx = D.load_archive_index(bytes(rom))
    b1_addr = D.IDX_BASE + idx[BLOCK1_ARCHIVE_IDX]
    b1_size = (D.IDX_BASE + idx[BLOCK1_ARCHIVE_IDX + 1]) - b1_addr
    block1 = P.parse_dialogue_block(bytes(rom), b1_addr, b1_size, dic)

    # 턴을 순서대로 평탄화
    flat_turns = []     # (dlg_idx, turn) 목록
    for dlg in block1['dialogues']:
        for turn in dlg['turns']:
            flat_turns.append((dlg['idx'], turn))
    n_dialogues = len(block1['dialogues'])
    print(f'  블록1: 0x{b1_addr:X}, {b1_size} 바이트, '
          f'대사 {n_dialogues}개, 턴 {len(flat_turns)}개')

    # ── 2. 매칭 엑셀에서 한국어 읽기 ──────────────────────
    from openpyxl import load_workbook
    ws = load_workbook(excel_path, data_only=True)['매칭 결과']
    excel_rows = []
    for r in range(2, ws.max_row + 1):
        cid_s = ws.cell(r, 3).value
        kr    = ws.cell(r, 8).value
        excel_rows.append((cid_s, kr))
    print(f'  엑셀 : {len(excel_rows)} 행')

    if len(excel_rows) != len(flat_turns):
        print(f'  ★주의: 엑셀 행({len(excel_rows)})과 ROM 턴({len(flat_turns)}) '
              f'수가 다릅니다. 순서 매칭에 오차가 있을 수 있습니다.')

    # cid 일치 검사
    mismatch = 0
    for i, (dlg_idx, turn) in enumerate(flat_turns):
        if i < len(excel_rows):
            cid_s = excel_rows[i][0]
            if cid_s:
                try:
                    if int(str(cid_s), 0) != turn['cid']:
                        mismatch += 1
                except ValueError:
                    pass
    if mismatch:
        print(f'  ★주의: 화자ID 불일치 턴 {mismatch}개 (엑셀↔ROM 정렬 점검 필요)')

    # ── 3. 코덱 준비 & 코드 배정 ─────────────────────────
    codec = HangulCodec(bytes(rom), korea_path, japan_path,
                        os.path.join(here, 'seg1_victim_rank.json'))
    all_kr = [kr for _, kr in excel_rows if kr]
    codec.plan(all_kr)
    print()
    print('  [코드 배정 결과]')
    print(codec.report())

    if codec.unresolved:
        print('  ※ 미배정 문자는 인코딩 시 생략됩니다.')

    # ── 4. 줄바꿈 + 인코딩 + 턴 재구성 ───────────────────
    print()
    print('  [턴별 인코딩]')
    new_turn_bytes = [None] * len(flat_turns)
    warn_turns = []
    kept_jp = 0

    for i, (dlg_idx, turn) in enumerate(flat_turns):
        kr = excel_rows[i][1] if i < len(excel_rows) else None
        jp_lines = len(turn['lines'])

        if not kr or not str(kr).strip():
            # 번역 없음 → 원본 일본어 유지
            new_turn_bytes[i] = rebuild_turn_original(turn)
            kept_jp += 1
            continue

        lines, warns = fit_turn_lines(str(kr), jp_lines, args.reserve)
        try:
            new_turn_bytes[i] = rebuild_turn_korean(turn, lines, codec)
        except ValueError as e:
            print(f'   턴#{i}: 인코딩 오류 → 원본 유지 ({e})')
            new_turn_bytes[i] = rebuild_turn_original(turn)
            kept_jp += 1
            continue
        if warns:
            extra = len(lines) - jp_lines     # 줄 수 초과량
            warn_turns.append((i, dlg_idx, jp_lines, len(lines), extra, warns))

    print(f'   한국어 인코딩 턴 : {len(flat_turns) - kept_jp}')
    print(f'   원본 일본어 유지 : {kept_jp} (번역 없음/오류)')
    if warn_turns:
        minor = [w for w in warn_turns if 0 < w[4] <= 1]
        major = [w for w in warn_turns if w[4] >= 2]
        width_only = [w for w in warn_turns if w[4] <= 0]
        print(f'   ⚠ 줄 수 +1 (한국어가 길어 자연스러운 확장): {len(minor)}턴')
        print(f'   ⚠ 줄 수 +2 이상 (번역 매칭 점검 권장): {len(major)}턴')
        if width_only:
            print(f'   ⚠ 폭만 초과: {len(width_only)}턴')
        if major:
            print('   [+2 이상 확장 턴 — 매칭 엑셀 점검 대상]')
            for i, di, jp, kr_n, ex, _ in major[:20]:
                print(f'      턴#{i}(대사{di}): 목표 {jp}줄 → 실제 {kr_n}줄')
            if len(major) > 20:
                print(f'      ... 외 {len(major) - 20}턴')

    # ── 5. 새 블록1 빌드 ─────────────────────────────────
    # 대사별로 턴 바이트 합치기
    dlg_bytes = []
    ti = 0
    for dlg in block1['dialogues']:
        buf = bytearray()
        for _ in dlg['turns']:
            buf += new_turn_bytes[ti]
            ti += 1
        dlg_bytes.append(bytes(buf))

    n_ptr = n_dialogues + 1
    table_size = n_ptr * 4
    ptrs = [table_size]
    for db in dlg_bytes:
        ptrs.append(ptrs[-1] + len(db))

    new_block = bytearray()
    for p in ptrs:
        new_block += struct.pack('<I', p)
    for db in dlg_bytes:
        new_block += db
    new_block = bytes(new_block)

    print()
    print(f'  [새 블록1] 크기 {len(new_block)} 바이트 '
          f'(원본 {b1_size} → {len(new_block):+d})')

    # ── 6. ROM 패치 ──────────────────────────────────────
    new_addr = (args.addr + 3) & ~3            # 4바이트 정렬
    if new_addr + len(new_block) > len(rom):
        sys.exit('[오류] 빈 공간이 부족합니다.')

    # 6-a) 새 블록 기록
    rom[new_addr:new_addr + len(new_block)] = new_block

    # 6-b) 아카이브 인덱스 repoint  (idx[1] → 새 위치)
    new_rel = new_addr - D.IDX_BASE
    idx_entry_off = D.IDX_BASE + BLOCK1_ARCHIVE_IDX * 4
    old_rel, = struct.unpack_from('<I', rom, idx_entry_off)
    struct.pack_into('<I', rom, idx_entry_off, new_rel)

    # 6-c) 사전 패치 (seg1 희생슬롯 덮어쓰기)
    for off, sjis in codec.dict_patches:
        rom[off:off + 2] = sjis

    print(f'  새 블록 기록 위치 : 0x{new_addr:X}')
    print(f'  인덱스[1] 수정    : 0x{old_rel:X} → 0x{new_rel:X}')
    print(f'  사전 패치         : {len(codec.dict_patches)} 슬롯')

    # ── 7. 저장 ─────────────────────────────────────────
    with open(out_path, 'wb') as f:
        f.write(rom)
    print()
    print(f'  ✓ 저장 완료: {out_path}')

    # ── 8. 검증: 다시 읽어 디코딩 ───────────────────────
    print()
    print('  [검증] 패치 ROM 재파싱')
    verify(out_path, new_addr, codec)


# ──────────────────────────────────────────────────────────
#  검증
# ──────────────────────────────────────────────────────────
def verify(rom_path, block_addr, codec):
    rom = D.load_rom(rom_path)
    dic = D.Dictionary(rom)

    # 인덱스가 새 위치를 가리키는지
    idx = D.load_archive_index(rom)
    resolved = D.IDX_BASE + idx[BLOCK1_ARCHIVE_IDX]
    if resolved != block_addr:
        print(f'   ★ 인덱스 검증 실패: 0x{resolved:X} != 0x{block_addr:X}')
        return
    print(f'   인덱스[1] → 0x{resolved:X}  OK')

    # 한자 → 한글 역맵 (출력용)
    kanji2ko = {v: k for k, v in codec.ko2kanji.items()}

    ptrs = P.read_dialogue_pointers(rom, block_addr)
    block_size = ptrs[-1]
    info = P.parse_dialogue_block(rom, block_addr, block_size, dic)
    ok_turns = sum(1 for d in info['dialogues'] for t in d['turns'])
    all_ok = all(d['parse_ok'] for d in info['dialogues'])
    print(f'   재파싱: 대사 {len(info["dialogues"])}개, 턴 {ok_turns}개, '
          f'파싱성공={"예" if all_ok else "★아니오"}')

    # 앞쪽 대사 몇 개를 한글로 복원해 출력
    print()
    print('   [한글 복원 미리보기]')
    shown = 0
    for dlg in info['dialogues']:
        for turn in dlg['turns']:
            txt_lines = []
            for line in turn['lines']:
                s = ''
                for code, w in D.tokenize(line):
                    ch = dic.decode(code)
                    # 디코딩 결과가 한자면 한글로 환원
                    s += ''.join(kanji2ko.get(c, c) for c in ch)
                txt_lines.append(s)
            preview = ' / '.join(txt_lines)
            print(f'    턴 cid=0x{turn["cid"]:02X}: {preview}')
            shown += 1
            if shown >= 8:
                print('    ...')
                return


if __name__ == '__main__':
    main()
